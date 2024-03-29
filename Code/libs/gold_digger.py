from libs.messages import Messages
from fastapi import Depends, FastAPI, HTTPException, status, Form, UploadFile
import datetime
from docx import Document
import pandas as pd
from unidecode import unidecode
import re
from io import BytesIO
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import threading

class GoldDigger:
    def __init__(self):
        self.__messages__ = Messages()
        self.__max_length__ = 1000
        self.__timeout_duration__ = 60
        self.__sup750__ = '>750 mil'
        self.__k750__ = '750 mil'
        self.__k585__ = '585 mil'
        self.__k375__ = '375 mil'
        self.__prix_bas__ = 'prix (€) bas'
        self.__prix_haut__ = 'prix (€) haut'
        
    async def docx_table_to_df(self, upload_file: UploadFile, table_index=0):
        # Read the content of the uploaded file into a BytesIO object
        content = await upload_file.read()
        file_stream = BytesIO(content)

        # Load the document
        doc = Document(file_stream)

        # Access the specific table
        table = doc.tables[table_index]

        # Extract and clean data from the table
        data = [[self.clean_text(cell.text) for cell in row.cells] for row in table.rows]

        # Create a DataFrame
        df = pd.DataFrame(data)

        # Optionally, set the first row as the header
        df.columns = df.iloc[0]
        df = df.drop(0)

        return df

    def clean_text(self, text):
        """ Transliterate non-UTF-8 characters to their closest UTF-8 equivalents """
        return unidecode(text)

    def extract_weights2(self, description):

        class TimeoutException(Exception):
            pass

        def timeout_handler():
            raise TimeoutException()

        # Dictionnaire pour stocker les poids
        weights = {self.__sup750__ : '', self.__k750__: '', self.__k585__: '', self.__k375__: ''}

        if len(description) > self.__max_length__ :
            raise ValueError("Entry too long")


        timer = threading.Timer(self.__timeout_duration__ , timeout_handler)
        try:
            timer.start()
            weight_parts = re.findall(
                r"(superieur a 750 mil|750 mil|585 mil|375 mil|Superieur a 750 mil) = (\d+\.?\d*) g", description)

        except TimeoutException:
            return "Timeout atteint pour l'expression régulière."
        finally:
            # Arrête le timer
            timer.cancel()

        # Parcourir les parties extraites et attribuer les poids aux catégories correspondantes
        for part in weight_parts:
            category, weight = part
            if category == 'superieur a 750 mil' or category == 'Superieur a 750 mil':
                weights[self.__sup750__ ] = weight
            elif category == self.__k750__:
                weights[self.__k750__] = weight
            elif category == self.__k585__:
                weights[self.__k585__] = weight
            elif category == self.__k375__:
                weights[self.__k375__] = weight

        return weights

    def extract_weight_and_separate_by_fineness(self, row):

        class TimeoutException(Exception):
            pass

        def timeout_handler():
            raise TimeoutException()

        # Initial setup for different fineness categories
        weights = {self.__sup750__ : None, self.__k750__: None, self.__k585__: None, self.__k375__: None}

        if len(row['Designation']) > self.__max_length__ :
            raise ValueError("L'entrée est trop longue.")

        # Regular expressions for finding weight and fineness
        weight_pattern = re.compile(r'(\d+(?:\.\d+)?)\s*g')
        fineness_pattern = re.compile(r'(\d{3,4})\s*mil', re.IGNORECASE)
        superior_pattern = re.compile(r'superieur a\s*(\d+)\s*mil', re.IGNORECASE)

        # Extracting weight
        # Crée un timer pour le timeout
        timer = threading.Timer(self.__timeout_duration__ , timeout_handler)
        try:
            timer.start()
            # Applique l'expression régulière
            weight_match = weight_pattern.search(row['Designation'])
        except TimeoutException:
            return "Timeout atteint pour l'expression régulière."
        finally:
            # Arrête le timer
            timer.cancel()

        if weight_match:
            weight = float(weight_match.group(1))

            # Determining fineness category and assigning weight
            fineness_match = fineness_pattern.search(row['Designation'])
            superior_match = superior_pattern.search(row['Designation'])
            if fineness_match and not superior_match:
                fineness = fineness_match.group(1) + ' mil'
                weights[fineness] = weight
            elif superior_match:
                fineness = ">" + superior_match.group(1) + ' mil'
                weights[fineness] = weight

        return pd.Series(weights)

    async def compute_excel_file(self, upload_file: UploadFile, price_per_kg: int, gold_coeffs: dict, output_file: str):

        price_per_g = price_per_kg / 1000

        # Extract the table
        df = await self.docx_table_to_df(upload_file)

        df['Designation'] = df['Designation'].str.replace(',', '.')

        df['Platine'] = df['Designation'].apply(lambda x: 'x' if 'platine' in x.lower() else "")
        # Créer des colonnes pour chaque titrage dans le DataFrame
        for mil in [self.__sup750__ , self.__k750__, self.__k585__, self.__k375__]:
            df[mil] = None

        # Appliquer la fonction d'extraction à chaque ligne
        for index, row in df.iterrows():
            weight_info = self.extract_weights2(row['Designation'])
            for key in weight_info:
                df.at[index, key] = weight_info[key]

        mask = df[[self.__sup750__ , self.__k750__, self.__k585__, self.__k375__]].isna() | (
                    df[[self.__sup750__ , self.__k750__, self.__k585__, self.__k375__]] == '')
        # Example of usage
        # Assuming 'data' is your DataFrame loaded from the Excel file

        mask = mask.all(axis=1)

        new_columns = df[mask].apply(self.extract_weight_and_separate_by_fineness, axis=1)
        df.update(new_columns)

        # Supposons que df est votre DataFrame
        # Convertissez d'abord les chaînes vides en NaN
        df.replace('', np.nan, inplace=True)

        # Ensuite, remplissez toutes les valeurs NaN par 0.0
        df.fillna(0.0, inplace=True)

        df[self.__k585__] = pd.to_numeric(df[self.__k585__], errors='coerce')
        df[self.__k375__] = pd.to_numeric(df[self.__k375__], errors='coerce')
        df[self.__k750__] = pd.to_numeric(df[self.__k750__], errors='coerce')
        df[self.__sup750__ ] = pd.to_numeric(df[self.__sup750__ ], errors='coerce')

        # Après la conversion, utilisez fillna pour remplacer les NaN par 0.0 si nécessaire
        df[self.__k585__].fillna(0.0, inplace=True)
        df[self.__k375__].fillna(0.0, inplace=True)
        df[self.__k750__].fillna(0.0, inplace=True)
        df[self.__sup750__ ].fillna(0.0, inplace=True)

        df[self.__prix_haut__] = ((price_per_g - gold_coeffs['offset_euros']/1000) * \
                              ( df[self.__k585__] * gold_coeffs['coeff_585_nume']/gold_coeffs['coeff_585_nume']
                                + df[self.__k375__] * gold_coeffs['coeff_375_nume']/gold_coeffs['coeff_375_nume']
                                + df[self.__k750__] * gold_coeffs['coeff_750_nume']/gold_coeffs['coeff_750_nume']
                                + df[self.__sup750__ ] * gold_coeffs['coeff_22up_nume']/gold_coeffs['coeff_22up_denum'])).round(0)

        df[self.__prix_bas__] = ((price_per_g - gold_coeffs['offset_euros']/1000) * \
                              ( df[self.__k585__] * gold_coeffs['coeff_585_nume']/gold_coeffs['coeff_585_nume']
                                + df[self.__k375__] * gold_coeffs['coeff_375_nume']/gold_coeffs['coeff_375_nume']
                                + df[self.__k750__] * gold_coeffs['coeff_750_nume']/gold_coeffs['coeff_750_nume']
                                + df[self.__sup750__ ] * gold_coeffs['coeff_22down_nume']/gold_coeffs['coeff_22down_denum'])).round(0)

        df[self.__prix_bas__] = df[self.__prix_bas__].astype(object)
        df[self.__prix_haut__] = df[self.__prix_haut__].astype(object)
        df.loc[df['Platine'] == 'x', self.__prix_haut__] = 'Platine'
        df.loc[df['Platine'] == 'x', self.__prix_bas__] = 'Platine'
        df.loc[df['Platine'] == 0, 'Platine'] = ''
        file_name = './data_out/' + datetime.datetime.now().strftime(
            "%Y%m%d%H%M%S") + '_mon_fichier_excel.xlsx'

        # Écrivez votre DataFrame dans le fichier Excel
        df.to_excel(output_file, index=False, sheet_name='Sheet1')

        # Chargez le fichier Excel pour la mise en forme avec openpyxl
        book = load_workbook(output_file)
        sheet = book['Sheet1']

        # Ajustement de la largeur des colonnes autres que 'Designation'
        for col in sheet.columns:
            if col[0].column_letter != 'B':  # 'B' est la lettre de la colonne pour 'Designation'
                max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
                sheet.column_dimensions[col[0].column_letter].width = max_length

        # Définir une largeur fixe pour la colonne 'Designation' et activer le texte à la ligne
        sheet.column_dimensions['B'].width = 70  # Ajustez la largeur selon vos besoins
        for cell in sheet['B']:
            cell.alignment = Alignment(wrap_text=True)

        # Ajuster la hauteur des lignes pour accommoder le texte mis à la ligne automatiquement
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].height = 70  # Ajustez la hauteur selon vos besoins

        # Créez des objets d'alignement pour chaque type de cellule
        alignment_title = Alignment(horizontal='center', vertical='top', wrap_text=True)
        alignment_designation = Alignment(horizontal='left', vertical='top', wrap_text=True)
        alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Appliquez l'alignement au titre
        for cell in sheet[1]:  # La première ligne contient les titres
            cell.alignment = alignment_title

        # Appliquez l'alignement à la colonne 'Designation'
        for cell in sheet['B']:
            cell.alignment = alignment_designation

        # Appliquez l'alignement centré à toutes les autres cellules
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                if cell.column_letter != 'B':  # Ne pas réappliquer pour la colonne 'Designation'
                    cell.alignment = alignment_center

        # Sauvegardez le fichier après la mise en forme
        book.save(output_file)
        book.close()

        return output_file
